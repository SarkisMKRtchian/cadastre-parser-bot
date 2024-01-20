from cd_parser import ParseExcel

from openpyxl import load_workbook, styles
from telebot import types, TeleBot
import time

import pathlib
import pandas
import os
import re

import log

class Excel:
    """
    Класс для обработки excel файла
    """
    def __init__(self, bot: TeleBot, parser: ParseExcel) -> None:
        self.bot = bot
        self.parser = parser
        
    def read(self, filePath: str, message: types.Message):
        """
        Открывает excel файл и забирает из него след. информацию
        Из Раздела 1
        КН здания
        Номер строки и колонки на котором расположен КН и где надо будет вставить обработаный текст
        КН зем. участка если имеется
        Из Раздела 2 
        Во первых удаляет все ключи, потому что они могут отличатся
        Далее проверят заполнены ли какие-то данные в яйчейке где должен лежать обработаный текст, если да то забирает отуда текст а если нет то оставляет пустую строку(Раздел 1 так же)
        После собраный обьект передает парсеру для обработки
        """
        try:
            sheet1 = pandas.read_excel(filePath, sheet_name='Раздел 1').to_dict()
            sheet2Dict = pandas.read_excel(filePath, sheet_name='Раздел 7').to_dict()
            
            sheet2 = []
            for key in sheet2Dict:
                sheet2.append(sheet2Dict[key])
            
            cadNums = [{
                    "col": 3, 
                    "row": 2, 
                    "sheet": "Раздел 1", 
                    "cadNum": sheet1['Unnamed: 1'][0], 
                    "mess": sheet1['Unnamed: 2'][0] if (len(sheet1) >= 3) else '', 
                    "adress": sheet1['Unnamed: 1'][4]
                }]
            
            
            if(sheet1['Unnamed: 1'][13] != "данные отсутствуют"): 
                cadNums.append({
                        "col": 3, 
                        "row": 15, 
                        "sheet": "Раздел 1", 
                        "cadNum": sheet1['Unnamed: 1'][13], 
                        "mess": sheet1['Unnamed: 2'][13] if (len(sheet1) >= 3) else ''
                    })
            
            for num in sheet2[1]:        
                cadNums.append({
                    "col": 8,
                    "row": num + 2,
                    "sheet": "Раздел 7",
                    "cadNum": sheet2[1][num],
                    "mess": str(sheet2[8][num]) if (len(sheet2) == 9) else ''
                })
            
            self.parser.parse(cadNums, filePath, message, Excel(self.bot, self.parser))
            
        except KeyError as ex:
            log.write(f"Key error: {ex} | {__file__}")
            self.bot.send_message(message.chat.id, "Ошибка при обработке файла. Провертье корректность файла")
            os.remove(filePath)
            
    def write(self, filePath: str, data: dict, chatId: int):
        """
        Пишет xls файл
        """
        try:
            wb = load_workbook(filePath, data_only=True)
            
            for item in data['data']:
                ws = wb[item["sheet"]]
                if(item['sheet'] == 'Раздел 1'):
                    ws.cell(item['row'], item['col']).value = item['mess']
                    ws.cell(item['row'], item['col']).alignment = styles.Alignment(horizontal="left", vertical="top")
                else:
                    ws.cell(item['row'], item['col']).value = data['data'][0]['cadNum']
                    ws.cell(item['row'], item['col'] + 1).value = item['mess']
                    
                    ws.cell(item['row'], item['col']).alignment = styles.Alignment(horizontal="left", vertical="top")
                    ws.cell(item['row'], item['col'] + 1).alignment = styles.Alignment(horizontal="left", vertical="top")
            
            adress: str = re.sub(r"[,|.| ]", '-', data['data'][0]['adress'])
            cad_num: str = re.sub(r"[:]", "-", data['data'][0]['cadNum'])
            fileName = f"Р1Р7-{cad_num}-{adress}{pathlib.Path(filePath).suffix}"
            
            wb.save(fileName)
            wb.close()
            os.remove(filePath)
            
            errs = ''
            if(len(data['processedFailure']) > 0):
                errs += 'При обработке следующих кад номеров произошла ошибка:\n'
                for cn in data['processedFailure']:
                    errs += f"{cn}\n"
                    
            stop = ''
            if(data['stoped']):
                stop += "Обработка остановлена\n"
                
            with open(fileName, "rb") as f:
                self.bot.send_document(chatId, f, caption=f"{stop}{data['data'][0]['adress']}\nC {time.strftime('%H:%M:%S', time.localtime(data['start']))} по {time.strftime('%H:%M:%S', time.localtime(data['end']))} = {round(data['start'] - data['end'], 1)}сек\n{data['processed']} (1кн: {data['timeForOneCard']}c.)\n{errs}")
                f.close()
            
            os.remove(fileName)
        except Exception as err:
            log.write(f"{err} | {__file__}")
            self.bot.send_message(chatId, "Не удалось сформировать exel документ. Повторитье попытку")